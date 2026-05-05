from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form, Response, Request, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
from passlib.context import CryptContext
import jwt
from collections import defaultdict
import pandas as pd
from io import BytesIO

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    raise ValueError("SECRET_KEY environment variable is required for security")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

security = HTTPBearer()

# Create the main app
app = FastAPI()
api_router = APIRouter(prefix="/api")

# Models
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    name: str
    role: str = "agent"  # admin, manager, agent
    team_code: Optional[str] = None
    manager_id: Optional[str] = None
    assigned_banks: Optional[List[str]] = None
    assigned_categories: Optional[List[str]] = None
    assigned_products: Optional[List[str]] = None
    active: bool = True  # Admin can enable/disable users
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    name: str
    password: str
    role: Optional[str] = "agent"
    team_code: Optional[str] = None
    manager_id: Optional[str] = None
    assigned_banks: Optional[List[str]] = None
    assigned_categories: Optional[List[str]] = None
    assigned_products: Optional[List[str]] = None

class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    team_code: Optional[str] = None
    manager_id: Optional[str] = None
    assigned_banks: Optional[List[str]] = None
    assigned_categories: Optional[List[str]] = None
    assigned_products: Optional[List[str]] = None
    active: Optional[bool] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str
    user: User

class CustomFieldConfig(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str  # Internal field name (e.g., "mobile_number")
    label: str  # Display label (e.g., "Mobile Number")
    field_type: str  # text, number, date, dropdown
    required: bool = False
    options: Optional[List[str]] = None  # For dropdown fields
    order: int = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomFieldConfigCreate(BaseModel):
    name: str
    label: str
    field_type: str
    required: bool = False
    options: Optional[List[str]] = None
    order: Optional[int] = 0

class CustomFieldConfigUpdate(BaseModel):
    label: Optional[str] = None
    field_type: Optional[str] = None
    required: Optional[bool] = None
    options: Optional[List[str]] = None
    order: Optional[int] = None


class Scheme(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class SchemeCreate(BaseModel):
    name: str
    description: Optional[str] = ""

class SchemeUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None


class Status(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = ""
    color: Optional[str] = "#3B82F6"  # Default blue color
    order: Optional[int] = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class StatusCreate(BaseModel):
    name: str
    description: Optional[str] = ""
    color: Optional[str] = "#3B82F6"
    order: Optional[int] = 0

class StatusUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    color: Optional[str] = None
    order: Optional[int] = None



class OrganizationSchedule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = "default"  # Only one schedule config
    working_days: List[str] = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    start_time: str = "09:00"
    end_time: str = "18:00"
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None

class OrganizationScheduleUpdate(BaseModel):
    working_days: Optional[List[str]] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None

class LoanApplication(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    agent_name: str
    customer_name: str
    company_name: str
    contact_no: str
    login_date: Optional[str] = ""
    status: str
    amount: Optional[str] = ""
    bank: str
    sanction: Optional[str] = ""
    disbursed: Optional[str] = ""
    disbursed_date: Optional[str] = ""
    remark: Optional[str] = ""
    decline_reason: Optional[str] = ""
    scheme: Optional[str] = ""
    case_from: Optional[str] = ""
    location: Optional[str] = ""
    branch: Optional[str] = ""
    executive_name: Optional[str] = ""
    team_manager: Optional[str] = ""
    code: Optional[str] = ""
    rate: Optional[str] = ""
    pf: Optional[str] = ""
    insurance: Optional[str] = ""
    tenure: Optional[str] = ""
    subvention: Optional[str] = ""
    brokerage_subvention: Optional[str] = ""
    category: Optional[str] = ""
    product: Optional[str] = ""
    technical_value: Optional[str] = ""
    legal_status: Optional[str] = ""
    month: str
    group_month: Optional[str] = ""
    custom_fields: Optional[Dict[str, Any]] = Field(default_factory=dict)
    entry_status: Optional[str] = "Open"  # "Open" or "Closed"
    file_count: Optional[int] = 0
    comment_count: Optional[int] = 0
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str

class LoanApplicationCreate(BaseModel):
    agent_name: str
    customer_name: str
    company_name: str
    contact_no: str
    login_date: Optional[str] = ""
    status: str
    amount: Optional[str] = ""
    bank: str
    sanction: Optional[str] = ""
    disbursed: Optional[str] = ""
    disbursed_date: Optional[str] = ""
    remark: Optional[str] = ""
    decline_reason: Optional[str] = ""
    scheme: Optional[str] = ""
    case_from: Optional[str] = ""
    location: Optional[str] = ""
    branch: Optional[str] = ""
    executive_name: Optional[str] = ""
    team_manager: Optional[str] = ""
    code: Optional[str] = ""
    rate: Optional[str] = ""
    pf: Optional[str] = ""
    insurance: Optional[str] = ""
    tenure: Optional[str] = ""
    subvention: Optional[str] = ""
    brokerage_subvention: Optional[str] = ""
    category: Optional[str] = ""
    product: Optional[str] = ""
    technical_value: Optional[str] = ""
    legal_status: Optional[str] = ""
    month: str
    group_month: Optional[str] = ""
    custom_fields: Optional[Dict[str, Any]] = None

class LoanApplicationUpdate(BaseModel):
    agent_name: Optional[str] = None
    customer_name: Optional[str] = None
    company_name: Optional[str] = None
    contact_no: Optional[str] = None
    login_date: Optional[str] = None
    status: Optional[str] = None
    amount: Optional[str] = None
    bank: Optional[str] = None
    sanction: Optional[str] = None
    disbursed: Optional[str] = None
    disbursed_date: Optional[str] = None
    remark: Optional[str] = None
    decline_reason: Optional[str] = None
    scheme: Optional[str] = None
    case_from: Optional[str] = None
    location: Optional[str] = None
    branch: Optional[str] = None
    executive_name: Optional[str] = None
    team_manager: Optional[str] = None
    code: Optional[str] = None
    rate: Optional[str] = None
    pf: Optional[str] = None
    insurance: Optional[str] = None
    tenure: Optional[str] = None
    subvention: Optional[str] = None
    brokerage_subvention: Optional[str] = None
    category: Optional[str] = None
    product: Optional[str] = None
    technical_value: Optional[str] = None
    legal_status: Optional[str] = None
    month: Optional[str] = None
    entry_status: Optional[str] = None
    group_month: Optional[str] = None
    custom_fields: Optional[Dict[str, Any]] = None

# Auth helpers
def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> User:
    token = credentials.credentials
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid authentication credentials")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token has expired")
    except jwt.JWTError:
        raise HTTPException(status_code=401, detail="Could not validate credentials")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return User(**user)

def check_admin(user: User):
    if user.role != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")

async def get_accessible_user_ids(user: User) -> List[str]:
    if user.role == "admin":
        return None
    elif user.role == "manager":
        team_users = await db.users.find(
            {"$or": [{"manager_id": user.id}, {"id": user.id}]},
            {"_id": 0, "id": 1}
        ).to_list(1000)
        return [u["id"] for u in team_users]
    else:
        return [user.id]

def build_rbac_filter(current_user: User, accessible_ids):
    """Build RBAC query filter. Agents always see their own loans. Managers see all team loans."""
    if current_user.role == 'admin' or accessible_ids is None:
        return {}
    
    # Agent: only own loans, no bank/cat/prod filtering needed
    # (agent always sees everything they created)
    if current_user.role == 'agent':
        return {"created_by": current_user.id}
    
    # Manager: see all team loans without bank/cat/prod restriction
    # Team visibility is controlled by manager_id assignment, not bank names
    return {"created_by": {"$in": accessible_ids}}

# Auth routes
@api_router.post("/auth/register", response_model=Token)
async def register(user_data: UserCreate):
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(
        email=user_data.email,
        name=user_data.name,
        role=user_data.role,
        team_code=user_data.team_code,
        manager_id=user_data.manager_id,
        assigned_banks=user_data.assigned_banks,
        assigned_categories=user_data.assigned_categories,
        assigned_products=user_data.assigned_products
    )
    
    user_dict = user.model_dump()
    user_dict['created_at'] = user_dict['created_at'].isoformat()
    user_dict['password'] = hash_password(user_data.password)
    
    await db.users.insert_one(user_dict)
    
    access_token = create_access_token(data={"sub": user.id})
    return Token(access_token=access_token, token_type="bearer", user=user)

@api_router.post("/auth/login", response_model=Token)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    if not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    # Check if user is active
    if not user.get('active', True):
        raise HTTPException(status_code=403, detail="Your account has been deactivated. Please contact admin.")
    
    if isinstance(user.get('created_at'), str):
        user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    user_obj = User(**{k: v for k, v in user.items() if k != 'password'})
    access_token = create_access_token(data={"sub": user_obj.id})
    return Token(access_token=access_token, token_type="bearer", user=user_obj)

@api_router.get("/auth/me", response_model=User)
async def get_me(current_user: User = Depends(get_current_user)):
    return current_user

# Change password route
class PasswordChange(BaseModel):
    old_password: str
    new_password: str

@api_router.post("/auth/change-password")
async def change_password(password_data: PasswordChange, current_user: User = Depends(get_current_user)):
    user = await db.users.find_one({"id": current_user.id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not verify_password(password_data.old_password, user['password']):
        raise HTTPException(status_code=400, detail="Current password is incorrect")
    
    new_hashed = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": current_user.id},
        {"$set": {"password": new_hashed}}
    )
    
    return {"message": "Password changed successfully"}

# User management (Admin only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    for user in users:
        if isinstance(user.get('created_at'), str):
            user['created_at'] = datetime.fromisoformat(user['created_at'])
    
    return users

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: UserUpdate, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    
    update_dict = {k: v for k, v in user_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.users.update_one({"id": user_id}, {"$set": update_dict})
    
    updated_user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not updated_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if isinstance(updated_user.get('created_at'), str):
        updated_user['created_at'] = datetime.fromisoformat(updated_user['created_at'])
    
    return User(**updated_user)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}


# Toggle user active status (Admin only)
@api_router.patch("/users/{user_id}/toggle-status")
async def toggle_user_status(user_id: str, current_user: User = Depends(get_current_user)):
    """Enable or disable a user (Admin only)"""
    check_admin(current_user)
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Toggle active status
    new_status = not user.get('active', True)
    
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"active": new_status}}
    )
    
    return {
        "message": f"User {'activated' if new_status else 'deactivated'} successfully",
        "active": new_status
    }


# Admin password reset
class PasswordResetRequest(BaseModel):
    new_password: str

@api_router.post("/users/{user_id}/reset-password")
async def admin_reset_password(
    user_id: str, 
    password_data: PasswordResetRequest, 
    current_user: User = Depends(get_current_user)
):
    """Admin can reset any user's password"""
    check_admin(current_user)
    
    # Check if user exists
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Hash and update password
    new_hashed = hash_password(password_data.new_password)
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"password": new_hashed}}
    )
    
    return {"message": f"Password reset successfully for {user.get('email', 'user')}"}


# Organization Schedule routes (Admin only)
@api_router.get("/organization/schedule", response_model=OrganizationSchedule)
async def get_organization_schedule(current_user: User = Depends(get_current_user)):
    """Get organization working schedule"""
    schedule = await db.organization_schedule.find_one({"id": "default"}, {"_id": 0})
    
    if not schedule:
        # Return default schedule if none exists
        default_schedule = OrganizationSchedule()
        return default_schedule
    
    if isinstance(schedule.get('updated_at'), str):
        schedule['updated_at'] = datetime.fromisoformat(schedule['updated_at'])
    
    return OrganizationSchedule(**schedule)

@api_router.put("/organization/schedule", response_model=OrganizationSchedule)
async def update_organization_schedule(
    schedule_data: OrganizationScheduleUpdate, 
    current_user: User = Depends(get_current_user)
):
    """Update organization working schedule (Admin only)"""
    check_admin(current_user)
    
    update_dict = {k: v for k, v in schedule_data.model_dump().items() if v is not None}
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    update_dict['updated_by'] = current_user.id
    
    # Check if schedule exists
    existing = await db.organization_schedule.find_one({"id": "default"})
    
    if existing:
        await db.organization_schedule.update_one(
            {"id": "default"},
            {"$set": update_dict}
        )
    else:
        # Create new schedule
        schedule = OrganizationSchedule(**schedule_data.model_dump(), updated_by=current_user.id)
        schedule_dict = schedule.model_dump()
        schedule_dict['updated_at'] = schedule_dict['updated_at'].isoformat()
        await db.organization_schedule.insert_one(schedule_dict)
    
    updated_schedule = await db.organization_schedule.find_one({"id": "default"}, {"_id": 0})
    
    if isinstance(updated_schedule.get('updated_at'), str):
        updated_schedule['updated_at'] = datetime.fromisoformat(updated_schedule['updated_at'])
    
    return OrganizationSchedule(**updated_schedule)

# Custom Field Configuration routes (Admin only)
@api_router.get("/field-configs", response_model=List[CustomFieldConfig])
async def get_field_configs(current_user: User = Depends(get_current_user)):
    """Get all custom field configurations"""
    fields = await db.field_configs.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    
    for field in fields:
        if isinstance(field.get('created_at'), str):
            field['created_at'] = datetime.fromisoformat(field['created_at'])
    
    return fields

@api_router.post("/field-configs", response_model=CustomFieldConfig)
async def create_field_config(field_data: CustomFieldConfigCreate, current_user: User = Depends(get_current_user)):
    """Create a new custom field configuration (Admin only)"""
    check_admin(current_user)
    
    # Check if field name already exists
    existing = await db.field_configs.find_one({"name": field_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Field name already exists")
    
    field = CustomFieldConfig(**field_data.model_dump())
    
    field_dict = field.model_dump()
    field_dict['created_at'] = field_dict['created_at'].isoformat()
    
    await db.field_configs.insert_one(field_dict)
    return field

@api_router.put("/field-configs/{field_id}", response_model=CustomFieldConfig)
async def update_field_config(field_id: str, field_data: CustomFieldConfigUpdate, current_user: User = Depends(get_current_user)):
    """Update a custom field configuration (Admin only)"""
    check_admin(current_user)
    
    update_dict = {k: v for k, v in field_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.field_configs.update_one({"id": field_id}, {"$set": update_dict})
    
    updated_field = await db.field_configs.find_one({"id": field_id}, {"_id": 0})
    if not updated_field:
        raise HTTPException(status_code=404, detail="Field configuration not found")
    
    if isinstance(updated_field.get('created_at'), str):
        updated_field['created_at'] = datetime.fromisoformat(updated_field['created_at'])
    
    return CustomFieldConfig(**updated_field)

@api_router.delete("/field-configs/{field_id}")
async def delete_field_config(field_id: str, current_user: User = Depends(get_current_user)):
    """Delete a custom field configuration (Admin only)"""
    check_admin(current_user)
    
    result = await db.field_configs.delete_one({"id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Field configuration not found")
    
    return {"message": "Field configuration deleted successfully"}


# Scheme Management routes (Admin/Manager only)
@api_router.get("/schemes", response_model=List[Scheme])
async def get_schemes(current_user: User = Depends(get_current_user)):
    """Get all schemes"""
    schemes = await db.schemes.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    
    for scheme in schemes:
        if isinstance(scheme.get('created_at'), str):
            scheme['created_at'] = datetime.fromisoformat(scheme['created_at'])
    
    return schemes

@api_router.post("/schemes", response_model=Scheme)
async def create_scheme(scheme_data: SchemeCreate, current_user: User = Depends(get_current_user)):
    """Create a new scheme (Admin/Manager only)"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can create schemes")
    
    # Check if scheme name already exists
    existing = await db.schemes.find_one({"name": scheme_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Scheme name already exists")
    
    scheme = Scheme(**scheme_data.model_dump(), created_by=current_user.id)
    
    scheme_dict = scheme.model_dump()
    scheme_dict['created_at'] = scheme_dict['created_at'].isoformat()
    
    await db.schemes.insert_one(scheme_dict)
    return scheme

@api_router.put("/schemes/{scheme_id}", response_model=Scheme)
async def update_scheme(scheme_id: str, scheme_data: SchemeUpdate, current_user: User = Depends(get_current_user)):
    """Update a scheme (Admin/Manager only)"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can update schemes")
    
    update_dict = {k: v for k, v in scheme_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.schemes.update_one({"id": scheme_id}, {"$set": update_dict})
    
    updated_scheme = await db.schemes.find_one({"id": scheme_id}, {"_id": 0})
    if not updated_scheme:
        raise HTTPException(status_code=404, detail="Scheme not found")
    
    if isinstance(updated_scheme.get('created_at'), str):
        updated_scheme['created_at'] = datetime.fromisoformat(updated_scheme['created_at'])
    
    return Scheme(**updated_scheme)

@api_router.delete("/schemes/{scheme_id}")
async def delete_scheme(scheme_id: str, current_user: User = Depends(get_current_user)):
    """Delete a scheme (Admin/Manager only)"""
    if current_user.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can delete schemes")
    
    result = await db.schemes.delete_one({"id": scheme_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Scheme not found")
    
    return {"message": "Scheme deleted successfully"}


# Status Management routes (Admin only)
@api_router.get("/statuses", response_model=List[Status])
async def get_statuses(current_user: User = Depends(get_current_user)):
    """Get all statuses"""
    statuses = await db.statuses.find({}, {"_id": 0}).sort("order", 1).to_list(1000)
    
    for status in statuses:
        if isinstance(status.get('created_at'), str):
            status['created_at'] = datetime.fromisoformat(status['created_at'])
    
    return statuses

@api_router.post("/statuses", response_model=Status)
async def create_status(status_data: StatusCreate, current_user: User = Depends(get_current_user)):
    """Create a new status (Admin only)"""
    check_admin(current_user)
    
    # Check if status name already exists
    existing = await db.statuses.find_one({"name": status_data.name})
    if existing:
        raise HTTPException(status_code=400, detail="Status name already exists")
    
    status = Status(**status_data.model_dump(), created_by=current_user.id)
    
    status_dict = status.model_dump()
    status_dict['created_at'] = status_dict['created_at'].isoformat()
    
    await db.statuses.insert_one(status_dict)
    return status

@api_router.put("/statuses/{status_id}", response_model=Status)
async def update_status(status_id: str, status_data: StatusUpdate, current_user: User = Depends(get_current_user)):
    """Update a status (Admin only)"""
    check_admin(current_user)
    
    update_dict = {k: v for k, v in status_data.model_dump().items() if v is not None}
    
    if update_dict:
        await db.statuses.update_one({"id": status_id}, {"$set": update_dict})
    
    updated_status = await db.statuses.find_one({"id": status_id}, {"_id": 0})
    if not updated_status:
        raise HTTPException(status_code=404, detail="Status not found")
    
    if isinstance(updated_status.get('created_at'), str):
        updated_status['created_at'] = datetime.fromisoformat(updated_status['created_at'])
    
    return Status(**updated_status)

@api_router.delete("/statuses/{status_id}")
async def delete_status(status_id: str, current_user: User = Depends(get_current_user)):
    """Delete a status (Admin only)"""
    check_admin(current_user)
    
    result = await db.statuses.delete_one({"id": status_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Status not found")
    
    return {"message": "Status deleted successfully"}

@api_router.post("/statuses/rename-in-loans")
async def rename_status_in_loans(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Bulk rename a status in all loan applications (Admin only)"""
    check_admin(current_user)
    old_name = data.get("old_name", "").strip()
    new_name = data.get("new_name", "").strip()
    if not old_name or not new_name:
        raise HTTPException(status_code=400, detail="old_name and new_name are required")
    result = await db.loan_applications.update_many(
        {"status": old_name},
        {"$set": {"status": new_name, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"Renamed '{old_name}' to '{new_name}' in {result.modified_count} loans", "modified_count": result.modified_count}

@api_router.get("/statuses/usage-count")
async def get_status_usage_count(current_user: User = Depends(get_current_user)):
    """Get count of loans using each status value"""
    check_admin(current_user)
    pipeline = [
        {"$group": {"_id": "$status", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    results = await db.loan_applications.aggregate(pipeline).to_list(1000)
    return {r["_id"]: r["count"] for r in results if r["_id"]}

# Loan routes with role-based access
@api_router.get("/loans")
async def get_loans(
    status: Optional[str] = None,
    bank: Optional[str] = None,
    agent_name: Optional[str] = None,
    month: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    limit: int = 500,
    current_user: User = Depends(get_current_user)
):
    query = {}
    
    # Role-based access control with own-loan bypass
    accessible_ids = await get_accessible_user_ids(current_user)
    rbac_filter = build_rbac_filter(current_user, accessible_ids)
    query.update(rbac_filter)
    
    if status:
        query["status"] = status
    if bank:
        query["bank"] = bank
    if agent_name:
        query["agent_name"] = agent_name
    if month:
        query["month"] = month
    if search:
        search_conditions = [
            {"customer_name": {"$regex": search, "$options": "i"}},
            {"company_name": {"$regex": search, "$options": "i"}},
            {"contact_no": {"$regex": search, "$options": "i"}}
        ]
        if "$or" in query:
            # RBAC already has $or, combine with $and
            rbac_or = query.pop("$or")
            query["$and"] = [{"$or": rbac_or}, {"$or": search_conditions}]
        else:
            query["$or"] = search_conditions
    
    # Cap limit to prevent abuse
    limit = min(limit, 2000)
    skip = (page - 1) * limit
    
    total = await db.loan_applications.count_documents(query)
    loans = await db.loan_applications.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    
    for loan in loans:
        if isinstance(loan.get('created_at'), str):
            loan['created_at'] = datetime.fromisoformat(loan['created_at'])
        if isinstance(loan.get('updated_at'), str):
            loan['updated_at'] = datetime.fromisoformat(loan['updated_at'])
    
    return {
        "loans": loans,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit
    }

@api_router.post("/loans", response_model=LoanApplication)
async def create_loan(loan_data: LoanApplicationCreate, current_user: User = Depends(get_current_user)):
    loan = LoanApplication(**loan_data.model_dump(), created_by=current_user.id)
    
    loan_dict = loan.model_dump()
    loan_dict['created_at'] = loan_dict['created_at'].isoformat()
    loan_dict['updated_at'] = loan_dict['updated_at'].isoformat()
    
    await db.loan_applications.insert_one(loan_dict)
    return loan

async def check_loan_access(loan_id: str, user: User) -> dict:
    loan = await db.loan_applications.find_one({"id": loan_id}, {"_id": 0})
    if not loan:
        raise HTTPException(status_code=404, detail="Loan application not found")
    
    if user.role == "admin":
        return loan
    
    accessible_ids = await get_accessible_user_ids(user)
    if loan["created_by"] not in accessible_ids:
        raise HTTPException(status_code=403, detail="Access denied")
    
    return loan

@api_router.get("/loans/{loan_id}", response_model=LoanApplication)
async def get_loan(loan_id: str, current_user: User = Depends(get_current_user)):
    loan = await check_loan_access(loan_id, current_user)
    
    if isinstance(loan.get('created_at'), str):
        loan['created_at'] = datetime.fromisoformat(loan['created_at'])
    if isinstance(loan.get('updated_at'), str):
        loan['updated_at'] = datetime.fromisoformat(loan['updated_at'])
    
    return LoanApplication(**loan)

@api_router.put("/loans/{loan_id}", response_model=LoanApplication)
async def update_loan(loan_id: str, loan_data: LoanApplicationUpdate, current_user: User = Depends(get_current_user)):
    loan = await check_loan_access(loan_id, current_user)
    
    update_dict = {k: v for k, v in loan_data.model_dump().items() if v is not None}
    update_dict['updated_at'] = datetime.now(timezone.utc).isoformat()
    
    await db.loan_applications.update_one({"id": loan_id}, {"$set": update_dict})
    
    updated_loan = await db.loan_applications.find_one({"id": loan_id}, {"_id": 0})
    if isinstance(updated_loan.get('created_at'), str):
        updated_loan['created_at'] = datetime.fromisoformat(updated_loan['created_at'])
    if isinstance(updated_loan.get('updated_at'), str):
        updated_loan['updated_at'] = datetime.fromisoformat(updated_loan['updated_at'])
    
    return LoanApplication(**updated_loan)

@api_router.delete("/loans/{loan_id}")
async def delete_loan(loan_id: str, current_user: User = Depends(get_current_user)):
    # Only admins can delete loans
    check_admin(current_user)
    
    result = await db.loan_applications.delete_one({"id": loan_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Loan application not found")
    
    return {"message": "Loan application deleted successfully"}

@api_router.patch("/loans/{loan_id}/entry-status")
async def toggle_entry_status(loan_id: str, data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Toggle entry_status between Open and Closed"""
    new_status = data.get("entry_status", "")
    if new_status not in ("Open", "Closed"):
        raise HTTPException(status_code=400, detail="entry_status must be 'Open' or 'Closed'")
    result = await db.loan_applications.update_one(
        {"id": loan_id},
        {"$set": {"entry_status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Loan not found")
    return {"message": f"Entry status set to {new_status}", "entry_status": new_status}

@api_router.post("/loans/bulk-delete")
async def bulk_delete_loans(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="No loan IDs provided")
    result = await db.loan_applications.delete_many({"id": {"$in": ids}})
    return {"message": f"{result.deleted_count} loans deleted", "deleted_count": result.deleted_count}

@api_router.post("/loans/bulk-status")
async def bulk_update_status(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    ids = data.get("ids", [])
    new_status = data.get("status", "")
    if not ids or not new_status:
        raise HTTPException(status_code=400, detail="ids and status are required")
    result = await db.loan_applications.update_many(
        {"id": {"$in": ids}},
        {"$set": {"status": new_status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": f"{result.modified_count} loans updated to '{new_status}'", "modified_count": result.modified_count}



@api_router.post("/loans/delete-by-date")
async def delete_loans_by_date(date_str: str, current_user: User = Depends(get_current_user)):
    """Delete all loans from a specific date - Admin only"""
    check_admin(current_user)
    
    try:
        # Parse the date string (e.g., "16-10-2025" or "2025-10-16")
        from datetime import datetime
        
        # Try different date formats
        date_formats = ["%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"]
        target_date = None
        
        for fmt in date_formats:
            try:
                target_date = datetime.strptime(date_str, fmt).date()
                break
            except:
                continue
        
        if not target_date:
            raise HTTPException(status_code=400, detail="Invalid date format. Use DD-MM-YYYY or YYYY-MM-DD")
        
        # Find all loans created on this date
        start_datetime = datetime.combine(target_date, datetime.min.time())
        end_datetime = datetime.combine(target_date, datetime.max.time())
        
        # Delete loans where created_at is on the target date
        result = await db.loan_applications.delete_many({
            "created_at": {
                "$gte": start_datetime.isoformat(),
                "$lte": end_datetime.isoformat()
            }
        })
        
        return {
            "message": f"Deleted {result.deleted_count} entries from {date_str}",
            "deleted_count": result.deleted_count
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete by date error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete entries: {str(e)}")

@api_router.post("/loans/delete-all")
async def delete_all_loans(confirm: str, current_user: User = Depends(get_current_user)):
    """Delete ALL loan entries - Admin only - DANGER"""
    check_admin(current_user)
    
    if confirm != "DELETE_ALL_DATA":
        raise HTTPException(status_code=400, detail="Confirmation text does not match")
    
    try:
        result = await db.loan_applications.delete_many({})
        
        return {
            "message": f"ALL MIS DATA DELETED! Removed {result.deleted_count} entries",
            "deleted_count": result.deleted_count
        }
        
    except Exception as e:
        logger.error(f"Delete all error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete all entries: {str(e)}")


@api_router.post("/loans/carry-forward")
async def carry_forward_loans(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Carry forward non-Disbursed loans from previous month to a new month"""
    to_month_key = data.get("to_month_key", "")
    if not to_month_key:
        raise HTTPException(status_code=400, detail="to_month_key is required (e.g., Apr-2026)")
    
    try:
        import re
        MONTH_NAMES_CF = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        
        # Parse target month
        match = re.match(r'^([A-Za-z]{3})-(\d{2,4})$', to_month_key)
        if not match:
            raise HTTPException(status_code=400, detail="Invalid month format")
        
        month_name = match.group(1)
        year = match.group(2)
        if len(year) == 2:
            year = f"20{year}"
        mi = MONTH_NAMES_CF.index(month_name) if month_name in MONTH_NAMES_CF else -1
        if mi == -1:
            raise HTTPException(status_code=400, detail="Invalid month name")
        
        # Calculate previous month
        prev_mi = mi - 1
        prev_year = int(year)
        if prev_mi < 0:
            prev_mi = 11
            prev_year -= 1
        prev_month_key = f"{MONTH_NAMES_CF[prev_mi]}-{prev_year}"
        prev_mm = str(prev_mi + 1).zfill(2)
        prev_yr = str(prev_year)
        
        # Build query to find previous month's loans (both original and carry-forwarded)
        month_patterns = [
            {"month": {"$regex": f"^\\d{{2}}-{prev_mm}-{prev_yr}$"}},
            {"month": {"$regex": f"^{MONTH_NAMES_CF[prev_mi]}-{prev_yr}$", "$options": "i"}},
            {"month": {"$regex": f"^{MONTH_NAMES_CF[prev_mi]}-{prev_yr[2:]}$", "$options": "i"}},
            {"month": {"$regex": f"^{prev_mm}-{prev_yr}$"}},
            {"month": {"$regex": f"^{prev_yr}-{prev_mm}-\\d{{2}}$"}},
            {"group_month": prev_month_key},  # Loans carried into previous month
        ]
        
        # RBAC filter
        accessible_ids = await get_accessible_user_ids(current_user)
        rbac = build_rbac_filter(current_user, accessible_ids)
        
        query = {"$or": month_patterns, "status": {"$ne": "Disbursed"}, "entry_status": {"$ne": "Closed"}}
        if rbac:
            query = {"$and": [rbac, {"$or": month_patterns}, {"status": {"$ne": "Disbursed"}}, {"entry_status": {"$ne": "Closed"}}]}
        
        prev_loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
        
        if not prev_loans:
            return {"message": f"No non-Disbursed loans found in {prev_month_key} to carry forward", "carried_count": 0, "from_month": prev_month_key}
        
        # Create copies for the new month
        new_mm = str(mi + 1).zfill(2)
        new_date = f"01-{new_mm}-{year}"
        
        new_loans = []
        for loan in prev_loans:
            new_loan = {**loan}
            new_loan["id"] = str(uuid.uuid4())
            new_loan["group_month"] = to_month_key
            new_loan["created_at"] = datetime.now(timezone.utc).isoformat()
            new_loans.append(new_loan)
        
        if new_loans:
            await db.loan_applications.insert_many(new_loans)
        
        return {
            "message": f"Carried forward {len(new_loans)} loans from {prev_month_key} to {to_month_key}",
            "carried_count": len(new_loans),
            "from_month": prev_month_key,
            "to_month": to_month_key
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Carry forward error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to carry forward: {str(e)}")


@api_router.post("/loans/delete-month")
async def delete_month_group(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    """Delete all loans in a month group and archive them to deleted_month_backups"""
    check_admin(current_user)
    month_key = data.get("month_key", "")
    if not month_key:
        raise HTTPException(status_code=400, detail="month_key is required")
    
    try:
        import re
        MONTH_NAMES = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        
        # Parse month_key (e.g., "Apr-2026") to find matching loans
        match = re.match(r'^([A-Za-z]{3})-(\d{2,4})$', month_key)
        if not match:
            raise HTTPException(status_code=400, detail="Invalid month_key format. Use MMM-YYYY (e.g., Apr-2026)")
        
        month_name = match.group(1)
        year = match.group(2)
        if len(year) == 2:
            year = f"20{year}"
        mi = MONTH_NAMES.index(month_name) if month_name in MONTH_NAMES else -1
        if mi == -1:
            raise HTTPException(status_code=400, detail="Invalid month name")
        mm = str(mi + 1).zfill(2)
        
        # Build query patterns to match various date formats for this month
        # dd-mm-yyyy, MMM-YYYY, mm-yyyy etc.
        month_patterns = [
            {"month": {"$regex": f"^\\d{{2}}-{mm}-{year}$"}},  # dd-mm-yyyy
            {"month": {"$regex": f"^{month_name}-{year}$", "$options": "i"}},  # MMM-YYYY
            {"month": {"$regex": f"^{mm}-{year}$"}},  # mm-yyyy
            {"month": {"$regex": f"^{year}-{mm}-\\d{{2}}$"}},  # yyyy-mm-dd
            {"group_month": month_key},  # Carry-forwarded loans
        ]
        # Also match 2-digit year
        short_year = year[2:]
        month_patterns.append({"month": {"$regex": f"^{month_name}-{short_year}$", "$options": "i"}})
        month_patterns.append({"month": {"$regex": f"^\\d{{2}}-{mm}-{short_year}$"}})
        
        query = {"$or": month_patterns}
        
        # Fetch loans to archive
        loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
        
        if not loans:
            return {"message": f"No loans found for {month_key}", "deleted_count": 0, "archived": False}
        
        # Archive to deleted_month_backups collection
        archive_doc = {
            "id": str(uuid.uuid4()),
            "month_key": month_key,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": current_user.name or current_user.email,
            "loan_count": len(loans),
            "loans": loans
        }
        await db.deleted_month_backups.insert_one(archive_doc)
        
        # Delete the loans
        result = await db.loan_applications.delete_many(query)
        
        return {
            "message": f"Deleted {result.deleted_count} loans from {month_key} and archived to backup",
            "deleted_count": result.deleted_count,
            "archived": True,
            "archive_id": archive_doc["id"]
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Delete month error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete month: {str(e)}")

@api_router.get("/backup/archived-months")
async def get_archived_months(current_user: User = Depends(get_current_user)):
    """Get list of archived month backups"""
    check_admin(current_user)
    archives = await db.deleted_month_backups.find({}, {"_id": 0, "loans": 0}).sort("deleted_at", -1).to_list(100)
    return archives

@api_router.post("/backup/restore-month/{archive_id}")
async def restore_month_backup(archive_id: str, current_user: User = Depends(get_current_user)):
    """Restore an archived month backup"""
    check_admin(current_user)
    archive = await db.deleted_month_backups.find_one({"id": archive_id}, {"_id": 0})
    if not archive:
        raise HTTPException(status_code=404, detail="Archive not found")
    
    loans = archive.get("loans", [])
    if loans:
        await db.loan_applications.insert_many(loans)
    
    # Remove from archive
    await db.deleted_month_backups.delete_one({"id": archive_id})
    
    return {"message": f"Restored {len(loans)} loans from {archive.get('month_key', 'unknown')}", "restored_count": len(loans)}

@api_router.delete("/backup/archived-month/{archive_id}")
async def delete_archived_month(archive_id: str, current_user: User = Depends(get_current_user)):
    """Permanently delete an archived month backup"""
    check_admin(current_user)
    result = await db.deleted_month_backups.delete_one({"id": archive_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Archive not found")
    return {"message": "Archive permanently deleted"}

@api_router.get("/backup/export-month/{month_key}")
async def export_month_loans(month_key: str, current_user: User = Depends(get_current_user)):
    """Export loans for a specific month as Excel"""
    import re
    MONTH_NAMES_EXP = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    
    match = re.match(r'^([A-Za-z]{3})-(\d{2,4})$', month_key)
    if not match:
        raise HTTPException(status_code=400, detail="Invalid month_key format")
    
    month_name = match.group(1)
    year = match.group(2)
    if len(year) == 2:
        year = f"20{year}"
    mi = MONTH_NAMES_EXP.index(month_name) if month_name in MONTH_NAMES_EXP else -1
    if mi == -1:
        raise HTTPException(status_code=400, detail="Invalid month name")
    mm = str(mi + 1).zfill(2)
    
    # Build RBAC filter
    accessible_ids = await get_accessible_user_ids(current_user)
    rbac = build_rbac_filter(current_user, accessible_ids)
    
    month_patterns = [
        {"month": {"$regex": f"^\\d{{2}}-{mm}-{year}$"}},
        {"month": {"$regex": f"^{month_name}-{year}$", "$options": "i"}},
        {"month": {"$regex": f"^{mm}-{year}$"}},
        {"month": {"$regex": f"^{year}-{mm}-\\d{{2}}$"}},
        {"group_month": month_key},
    ]
    short_year = year[2:]
    month_patterns.append({"month": {"$regex": f"^{month_name}-{short_year}$", "$options": "i"}})
    
    query = {"$or": month_patterns}
    if rbac:
        query = {"$and": [rbac, {"$or": month_patterns}]}
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    df = pd.DataFrame(loans)
    column_config = [
        ('month', 'Date'), ('customer_name', 'Customer Name'), ('company_name', 'Company Name'),
        ('contact_no', 'Contact No'), ('bank', 'Bank'), ('category', 'Category'), ('product', 'Product'),
        ('login_date', 'Login Date'), ('status', 'Status'), ('amount', 'Amount'), ('entry_status', 'Entry Status'),
        ('sanction', 'Sanction Amount'), ('disbursed', 'Disbursed Amount'), ('disbursed_date', 'Disbursed Date'),
        ('remark', 'Remark'), ('decline_reason', 'Decline Reason'), ('scheme', 'Scheme'),
        ('case_from', 'Case From'), ('location', 'Location'), ('branch', 'Branch'),
        ('executive_name', 'Executive Name'), ('team_manager', 'Team Manager'), ('code', 'Code'),
        ('rate', 'Rate'), ('pf', 'PF'), ('insurance', 'Insurance'), ('tenure', 'Tenure'),
        ('subvention', 'Subvention'), ('brokerage_subvention', 'Brokerage'), ('agent_name', 'Agent Name'),
        ('technical_value', 'Technical Value'), ('legal_status', 'Legal Status'),
    ]
    
    if not df.empty:
        for c_key, _ in column_config:
            if c_key not in df.columns:
                df[c_key] = ""
        ordered_cols = [c[0] for c in column_config]
        df = df[ordered_cols]
        rename_map = {c[0]: c[1] for c in column_config}
        df.rename(columns=rename_map, inplace=True)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=month_key)
        worksheet = writer.sheets[month_key]
        from openpyxl.utils import get_column_letter
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if not df.empty else 0)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=loans_{month_key}.xlsx"}
    )



@api_router.post("/loans/normalize-months")
async def normalize_months(current_user: User = Depends(get_current_user)):
    """Normalize all month formats to 'Mon-YY' format - Admin only"""
    check_admin(current_user)
    
    try:
        import re
        
        # Mapping of month names
        month_map = {
            'january': 'Jan', 'february': 'Feb', 'march': 'Mar', 'april': 'Apr',
            'may': 'May', 'june': 'Jun', 'july': 'Jul', 'august': 'Aug',
            'september': 'Sep', 'october': 'Oct', 'november': 'Nov', 'december': 'Dec'
        }
        
        all_loans = await db.loan_applications.find({}).to_list(10000)
        updated_count = 0
        
        for loan in all_loans:
            original_month = loan.get('month', '').strip()
            if not original_month:
                continue
                
            normalized_month = None
            original_lower = original_month.lower()
            
            # Try different formats
            # Format: "November", "NOVEMBER"
            if original_lower in month_map:
                normalized_month = f"{month_map[original_lower]}-25"  # Default to 25 if no year
            
            # Format: "Nov 25", "NOV 25"
            match = re.match(r'(\w{3})\s+(\d{2})', original_month, re.IGNORECASE)
            if match:
                month_abbr = match.group(1).capitalize()
                year = match.group(2)
                normalized_month = f"{month_abbr}-{year}"
            
            # Format: "Nov-25", "NOV-25", "nov-2025", "NOV-2025"
            match = re.match(r'(\w{3})-(\d{2,4})', original_month, re.IGNORECASE)
            if match:
                month_abbr = match.group(1).capitalize()
                year = match.group(2)
                if len(year) == 4:
                    year = year[2:]  # Convert 2025 to 25
                normalized_month = f"{month_abbr}-{year}"
            
            # Format: "November 2025", "NOVEMBER 2025"
            match = re.match(r'(\w+)\s+(\d{4})', original_month, re.IGNORECASE)
            if match:
                month_name = match.group(1).lower()
                year = match.group(2)[2:]  # Get last 2 digits
                if month_name in month_map:
                    normalized_month = f"{month_map[month_name]}-{year}"
            
            # Update if normalized
            if normalized_month and normalized_month != original_month:
                await db.loan_applications.update_one(
                    {"id": loan['id']},
                    {"$set": {"month": normalized_month}}
                )
                updated_count += 1
                logger.info(f"Normalized: '{original_month}' -> '{normalized_month}'")
        
        return {
            "message": f"Successfully normalized {updated_count} month values",
            "updated_count": updated_count,
            "total_checked": len(all_loans)
        }
        
    except Exception as e:
        logger.error(f"Month normalization error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to normalize months: {str(e)}")


# Analytics routes
@api_router.get("/analytics/overview")
async def get_overview(current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    total = len(loans)
    status_counts = defaultdict(int)
    disbursed_count = 0
    declined_count = 0
    pending_count = 0
    total_sanction_amount = 0
    total_disbursed_amount = 0
    
    for loan in loans:
        status = loan.get('status', '')
        status_counts[status] += 1
        
        # Calculate amounts
        sanction_str = loan.get('sanction', '0')
        disbursed_str = loan.get('disbursed', '0')
        
        try:
            sanction_amt = float(str(sanction_str).replace(',', '')) if sanction_str else 0
            total_sanction_amount += sanction_amt
        except:
            pass
        
        try:
            disbursed_amt = float(str(disbursed_str).replace(',', '')) if disbursed_str else 0
            total_disbursed_amount += disbursed_amt
        except:
            pass
        
        if status == 'Disbursed':
            disbursed_count += 1
        elif status == 'Decline':
            declined_count += 1
        elif status in ['Hold', 'Login Done', 'Sent For Login', 'Pd To Be Done']:
            pending_count += 1
    
    return {
        "total": total,
        "disbursed": disbursed_count,
        "declined": declined_count,
        "pending": pending_count,
        "total_sanction_amount": total_sanction_amount,
        "total_disbursed_amount": total_disbursed_amount,
        "status_breakdown": dict(status_counts)
    }

@api_router.get("/analytics/monthly-trends")
async def get_monthly_trends(current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    MONTH_NAMES_T = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    
    def to_month_key(loan):
        gm = loan.get('group_month', '')
        if gm:
            return gm
        val = loan.get('month', '')
        if not val:
            return 'Unknown'
        import re
        if re.match(r'^[A-Za-z]{3}-\d{4}$', val):
            return val
        parts = val.split('-')
        if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12:
                return f"{MONTH_NAMES_T[mi]}-{parts[2]}"
        if len(parts) == 3 and len(parts[0]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12:
                return f"{MONTH_NAMES_T[mi]}-{parts[0]}"
        return val
    
    monthly_data = defaultdict(lambda: {
        "total": 0, "disbursed": 0, "declined": 0, "pending": 0,
        "sanction_amount": 0, "disbursed_amount": 0, "month": ""
    })
    
    for loan in loans:
        month = to_month_key(loan)
        status = loan.get('status', '')
        
        monthly_data[month]["month"] = month
        monthly_data[month]["total"] += 1
        
        san = 0
        dis = 0
        try:
            san = float(str(loan.get('sanction', '') or '0').replace(',', ''))
        except: pass
        try:
            dis = float(str(loan.get('disbursed', '') or '0').replace(',', ''))
        except: pass
        monthly_data[month]["sanction_amount"] += san
        monthly_data[month]["disbursed_amount"] += dis
        
        if status == 'Disbursed':
            monthly_data[month]["disbursed"] += 1
        elif status in ('Decline', 'Rejected'):
            monthly_data[month]["declined"] += 1
        else:
            monthly_data[month]["pending"] += 1
    
    def sort_key(item):
        import re
        m = re.match(r'^([A-Za-z]{3})-(\d{4})$', item.get('month', ''))
        if m and m.group(1) in MONTH_NAMES_T:
            return f"{m.group(2)}-{str(MONTH_NAMES_T.index(m.group(1))).zfill(2)}"
        return item.get('month', '')
    
    sorted_data = sorted(monthly_data.values(), key=sort_key)
    return sorted_data

@api_router.get("/analytics/by-bank")
async def get_by_bank(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    MONTH_NAMES_BK = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    def to_mk(loan):
        gm = loan.get('group_month', '')
        if gm: return gm
        val = loan.get('month', '')
        if not val: return 'Unknown'
        import re
        if re.match(r'^[A-Za-z]{3}-\d{4}$', val): return val
        parts = val.split('-')
        if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12: return f"{MONTH_NAMES_BK[mi]}-{parts[2]}"
        if len(parts) == 3 and len(parts[0]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12: return f"{MONTH_NAMES_BK[mi]}-{parts[0]}"
        return val
    
    if month and month != 'all':
        loans = [l for l in loans if to_mk(l) == month]
    
    bank_stats = defaultdict(lambda: {"total": 0, "disbursed": 0, "declined": 0})
    
    for loan in loans:
        bank = loan.get('bank', 'Unknown')
        status = loan.get('status', '')
        
        bank_stats[bank]["total"] += 1
        if status == 'Disbursed':
            bank_stats[bank]["disbursed"] += 1
        elif status in ('Decline', 'Rejected'):
            bank_stats[bank]["declined"] += 1
    
    return dict(bank_stats)

@api_router.get("/analytics/by-agent")
async def get_by_agent(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    MONTH_NAMES_AG = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    def to_mk(loan):
        gm = loan.get('group_month', '')
        if gm: return gm
        val = loan.get('month', '')
        if not val: return 'Unknown'
        import re
        if re.match(r'^[A-Za-z]{3}-\d{4}$', val): return val
        parts = val.split('-')
        if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12: return f"{MONTH_NAMES_AG[mi]}-{parts[2]}"
        if len(parts) == 3 and len(parts[0]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12: return f"{MONTH_NAMES_AG[mi]}-{parts[0]}"
        return val
    
    if month and month != 'all':
        loans = [l for l in loans if to_mk(l) == month]
    
    agent_stats = defaultdict(lambda: {"total": 0, "disbursed": 0, "declined": 0, "pending": 0})
    
    for loan in loans:
        agent = loan.get('agent_name', 'Unknown')
        status = loan.get('status', '')
        
        agent_stats[agent]["total"] += 1
        if status == 'Disbursed':
            agent_stats[agent]["disbursed"] += 1
        elif status in ('Decline', 'Rejected'):
            agent_stats[agent]["declined"] += 1
        else:
            agent_stats[agent]["pending"] += 1
    
    return dict(agent_stats)

@api_router.get("/analytics/by-month")
async def get_by_month(current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    month_stats = defaultdict(lambda: {"total": 0, "disbursed": 0, "declined": 0})
    
    for loan in loans:
        month = loan.get('month', 'Unknown')
        status = loan.get('status', '')
        
        month_stats[month]["total"] += 1
        if status == 'Disbursed':
            month_stats[month]["disbursed"] += 1
        elif status == 'Decline':
            month_stats[month]["declined"] += 1
    
    return dict(month_stats)

@api_router.get("/analytics/deep")
async def get_deep_analytics(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Deep analytics: category-wise, product-wise, status distribution, amounts"""
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    if month and month != 'all':
        MONTH_NAMES_DP = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
        def to_mk(loan):
            gm = loan.get('group_month', '')
            if gm: return gm
            val = loan.get('month', '')
            if not val: return 'Unknown'
            import re
            if re.match(r'^[A-Za-z]{3}-\d{4}$', val): return val
            parts = val.split('-')
            if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
                mi = int(parts[1]) - 1
                if 0 <= mi < 12: return f"{MONTH_NAMES_DP[mi]}-{parts[2]}"
            if len(parts) == 3 and len(parts[0]) == 4:
                mi = int(parts[1]) - 1
                if 0 <= mi < 12: return f"{MONTH_NAMES_DP[mi]}-{parts[0]}"
            return val
        loans = [l for l in loans if to_mk(l) == month]
    
    category_stats = defaultdict(lambda: {"total": 0, "disbursed": 0, "sanction_amt": 0, "disbursed_amt": 0})
    product_stats = defaultdict(lambda: {"total": 0, "disbursed": 0, "sanction_amt": 0, "disbursed_amt": 0})
    status_dist = defaultdict(int)
    entry_status_dist = {"Open": 0, "Closed": 0}
    agent_amounts = defaultdict(lambda: {"total": 0, "sanction_amt": 0, "disbursed_amt": 0, "disbursed_count": 0})
    
    total_sanction = 0
    total_disbursed = 0
    
    for loan in loans:
        cat = loan.get('category', '') or 'Uncategorized'
        prod = loan.get('product', '') or 'Unassigned'
        status = loan.get('status', 'Unknown')
        agent = loan.get('agent_name', 'Unknown')
        es = loan.get('entry_status', 'Open')
        
        san = 0
        dis = 0
        try: san = float(str(loan.get('sanction', '') or '0').replace(',', ''))
        except: pass
        try: dis = float(str(loan.get('disbursed', '') or '0').replace(',', ''))
        except: pass
        
        total_sanction += san
        total_disbursed += dis
        
        category_stats[cat]["total"] += 1
        category_stats[cat]["sanction_amt"] += san
        category_stats[cat]["disbursed_amt"] += dis
        if status == 'Disbursed':
            category_stats[cat]["disbursed"] += 1
        
        product_stats[prod]["total"] += 1
        product_stats[prod]["sanction_amt"] += san
        product_stats[prod]["disbursed_amt"] += dis
        if status == 'Disbursed':
            product_stats[prod]["disbursed"] += 1
        
        status_dist[status] += 1
        
        if es == 'Closed':
            entry_status_dist["Closed"] += 1
        else:
            entry_status_dist["Open"] += 1
        
        agent_amounts[agent]["total"] += 1
        agent_amounts[agent]["sanction_amt"] += san
        agent_amounts[agent]["disbursed_amt"] += dis
        if status == 'Disbursed':
            agent_amounts[agent]["disbursed_count"] += 1
    
    return {
        "category_stats": dict(category_stats),
        "product_stats": dict(product_stats),
        "status_distribution": dict(status_dist),
        "entry_status_distribution": entry_status_dist,
        "agent_amounts": dict(sorted(agent_amounts.items(), key=lambda x: -x[1]["disbursed_amt"])),
        "totals": {"total_loans": len(loans), "total_sanction": total_sanction, "total_disbursed": total_disbursed}
    }

@api_router.get("/analytics/unique-values")
async def get_unique_values(current_user: User = Depends(get_current_user)):
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    MONTH_NAMES_UV = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    
    def to_month_key_uv(loan):
        gm = loan.get('group_month', '')
        if gm:
            return gm
        val = loan.get('month', '')
        if not val:
            return None
        import re
        if re.match(r'^[A-Za-z]{3}-\d{4}$', val):
            return val
        parts = val.split('-')
        if len(parts) == 3 and len(parts[0]) <= 2 and len(parts[2]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12:
                return f"{MONTH_NAMES_UV[mi]}-{parts[2]}"
        if len(parts) == 3 and len(parts[0]) == 4:
            mi = int(parts[1]) - 1
            if 0 <= mi < 12:
                return f"{MONTH_NAMES_UV[mi]}-{parts[0]}"
        return val
    
    banks = set()
    statuses = set()
    agents = set()
    months = set()
    schemes = set()
    
    for loan in loans:
        if loan.get('bank'):
            banks.add(loan['bank'])
        if loan.get('status'):
            statuses.add(loan['status'])
        if loan.get('agent_name'):
            agents.add(loan['agent_name'])
        mk = to_month_key_uv(loan)
        if mk:
            months.add(mk)
        if loan.get('scheme'):
            schemes.add(loan['scheme'])
    
    def month_sort_key(m):
        import re
        match = re.match(r'^([A-Za-z]{3})-(\d{4})$', m)
        if match and match.group(1) in MONTH_NAMES_UV:
            return f"{match.group(2)}-{str(MONTH_NAMES_UV.index(match.group(1))).zfill(2)}"
        return m
    
    return {
        "banks": sorted(list(banks)),
        "statuses": sorted(list(statuses)),
        "agents": sorted(list(agents)),
        "months": sorted(list(months), key=month_sort_key),
        "schemes": sorted(list(schemes))
    }


@api_router.get("/analytics/team-leaderboard")
async def get_team_leaderboard(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    """Returns agent performance leaderboard for managers/admins"""
    query = {}
    accessible_ids = await get_accessible_user_ids(current_user)
    query.update(build_rbac_filter(current_user, accessible_ids))

    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(50000)

    # Load targets for the selected month
    target_query = {}
    if month:
        target_query["month"] = month
    targets_list = await db.agent_targets.find(target_query, {"_id": 0}).to_list(5000)
    targets_map = {}
    for t in targets_list:
        key = t.get("agent_name", "")
        if month:
            targets_map[key] = t.get("target_amount", 0)
        else:
            targets_map.setdefault(key, 0)
            targets_map[key] += t.get("target_amount", 0)

    agent_stats = {}
    for loan in loans:
        agent = loan.get("agent_name") or loan.get("created_by") or "Unknown"
        if agent not in agent_stats:
            agent_stats[agent] = {"agent_name": agent, "total_loans": 0, "sanction_amount": 0, "disbursed_amount": 0, "disbursed_count": 0, "pending_count": 0, "declined_count": 0}
        stats = agent_stats[agent]
        stats["total_loans"] += 1
        s = float(str(loan.get("sanction", "0") or "0").replace(",", "") or 0)
        d = float(str(loan.get("disbursed", "0") or "0").replace(",", "") or 0)
        stats["sanction_amount"] += s if not __import__('math').isnan(s) else 0
        stats["disbursed_amount"] += d if not __import__('math').isnan(d) else 0
        status = (loan.get("status") or "").lower()
        if "disburse" in status:
            stats["disbursed_count"] += 1
        elif "decline" in status or "reject" in status:
            stats["declined_count"] += 1
        elif "pending" in status or "login" in status or "process" in status:
            stats["pending_count"] += 1

    leaderboard = sorted(agent_stats.values(), key=lambda x: x["disbursed_amount"], reverse=True)
    for i, entry in enumerate(leaderboard):
        entry["rank"] = i + 1
        entry["conversion_rate"] = round((entry["disbursed_count"] / entry["total_loans"] * 100), 1) if entry["total_loans"] > 0 else 0
        target = targets_map.get(entry["agent_name"], 0)
        entry["target_amount"] = target
        entry["target_progress"] = round((entry["disbursed_amount"] / target * 100), 1) if target > 0 else 0

    return {"leaderboard": leaderboard, "total_agents": len(leaderboard), "total_loans": len(loans)}


# Agent Monthly Targets
@api_router.get("/targets")
async def get_targets(month: Optional[str] = None, current_user: User = Depends(get_current_user)):
    query = {}
    if month:
        query["month"] = month
    targets = await db.agent_targets.find(query, {"_id": 0}).to_list(5000)
    return targets

@api_router.post("/targets")
async def set_target(data: dict = Body(...), current_user: User = Depends(get_current_user)):
    if current_user.role not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Only admin/manager can set targets")
    agent_name = data.get("agent_name")
    month = data.get("month")
    target_amount = data.get("target_amount", 0)
    if not agent_name or not month:
        raise HTTPException(status_code=400, detail="agent_name and month are required")
    existing = await db.agent_targets.find_one({"agent_name": agent_name, "month": month})
    if existing:
        await db.agent_targets.update_one(
            {"agent_name": agent_name, "month": month},
            {"$set": {"target_amount": float(target_amount), "updated_by": current_user.id, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    else:
        await db.agent_targets.insert_one({
            "id": str(uuid.uuid4()), "agent_name": agent_name, "month": month,
            "target_amount": float(target_amount),
            "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()
        })
    return {"status": "ok", "agent_name": agent_name, "month": month, "target_amount": float(target_amount)}

@api_router.delete("/targets/{agent_name}/{month}")
async def delete_target(agent_name: str, month: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ['admin', 'manager']:
        raise HTTPException(status_code=403, detail="Only admin/manager can delete targets")
    await db.agent_targets.delete_one({"agent_name": agent_name, "month": month})
    return {"status": "deleted"}



# Excel Export (Admin only)
@api_router.get("/export/loans")
async def export_loans(
    month: Optional[str] = None,
    status: Optional[str] = None,
    bank: Optional[str] = None,
    current_user: User = Depends(get_current_user)
):
    # Restrict to admin only
    check_admin(current_user)
    
    query = {}
    # Admin can export all data
    accessible_ids = await get_accessible_user_ids(current_user)
    if accessible_ids is not None:
        query["created_by"] = {"$in": accessible_ids}
    
    if month:
        query["month"] = month
    if status:
        query["status"] = status
    if bank:
        query["bank"] = bank
    
    loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
    
    df = pd.DataFrame(loans)
    
    # Define columns with proper readable headers
    column_config = [
        ('month', 'Date'),
        ('customer_name', 'Customer Name'),
        ('company_name', 'Company Name'),
        ('contact_no', 'Contact No'),
        ('bank', 'Bank'),
        ('category', 'Category'),
        ('product', 'Product'),
        ('login_date', 'Login Date'),
        ('status', 'Status'),
        ('amount', 'Amount'),
        ('entry_status', 'Entry Status'),
        ('sanction', 'Sanction Amount'),
        ('disbursed', 'Disbursed Amount'),
        ('disbursed_date', 'Disbursed Date'),
        ('remark', 'Remark'),
        ('decline_reason', 'Decline Reason'),
        ('scheme', 'Scheme'),
        ('case_from', 'Case From'),
        ('location', 'Location'),
        ('branch', 'Branch'),
        ('executive_name', 'Executive Name'),
        ('team_manager', 'Team Manager'),
        ('code', 'Code'),
        ('rate', 'Rate'),
        ('pf', 'PF'),
        ('insurance', 'Insurance'),
        ('tenure', 'Tenure'),
        ('subvention', 'Subvention'),
        ('brokerage_subvention', 'Brokerage'),
        ('agent_name', 'Agent Name'),
        ('technical_value', 'Technical Value'),
        ('legal_status', 'Legal Status'),
    ]
    
    if not df.empty:
        for c_key, _ in column_config:
            if c_key not in df.columns:
                df[c_key] = ""
        ordered_cols = [c[0] for c in column_config]
        df = df[ordered_cols]
        rename_map = {c[0]: c[1] for c in column_config}
        df.rename(columns=rename_map, inplace=True)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Loans')
        
        # Auto-fit column widths
        worksheet = writer.sheets['Loans']
        from openpyxl.utils import get_column_letter
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(len(str(col)), df[col].astype(str).str.len().max() if not df.empty else 0)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
    
    output.seek(0)
    
    filename = f"loans_export_{month if month else 'all'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/backup/full-data")
async def backup_all_data(current_user: User = Depends(get_current_user)):
    """Export all data as Excel with proper column headers - Admin only"""
    check_admin(current_user)
    
    # Get all loans
    loans = await db.loan_applications.find({}, {"_id": 0}).to_list(10000)
    
    # Get all users (without passwords)
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Get all field configs
    field_configs = await db.field_configs.find({}, {"_id": 0}).to_list(100)
    
    # Loan column config
    loan_column_config = [
        ('month', 'Date'), ('customer_name', 'Customer Name'), ('company_name', 'Company Name'),
        ('contact_no', 'Contact No'), ('bank', 'Bank'), ('category', 'Category'), ('product', 'Product'),
        ('login_date', 'Login Date'), ('status', 'Status'), ('amount', 'Amount'), ('entry_status', 'Entry Status'),
        ('sanction', 'Sanction Amount'), ('disbursed', 'Disbursed Amount'), ('disbursed_date', 'Disbursed Date'),
        ('remark', 'Remark'), ('decline_reason', 'Decline Reason'), ('scheme', 'Scheme'),
        ('case_from', 'Case From'), ('location', 'Location'), ('branch', 'Branch'),
        ('executive_name', 'Executive Name'), ('team_manager', 'Team Manager'), ('code', 'Code'),
        ('rate', 'Rate'), ('pf', 'PF'), ('insurance', 'Insurance'), ('tenure', 'Tenure'),
        ('subvention', 'Subvention'), ('brokerage_subvention', 'Brokerage'), ('agent_name', 'Agent Name'),
        ('technical_value', 'Technical Value'), ('legal_status', 'Legal Status'),
        ('group_month', 'Group Month'),
    ]
    
    user_column_config = [
        ('name', 'Name'), ('email', 'Email'), ('role', 'Role'), ('team_code', 'Team Code'),
        ('assigned_banks', 'Assigned Banks'), ('assigned_categories', 'Assigned Categories'),
        ('assigned_products', 'Assigned Products'), ('active', 'Active'),
    ]
    
    output = BytesIO()
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Loans sheet with proper columns
        if loans:
            df_loans = pd.DataFrame(loans)
            for c_key, _ in loan_column_config:
                if c_key not in df_loans.columns:
                    df_loans[c_key] = ""
            ordered_cols = [c[0] for c in loan_column_config]
            df_loans = df_loans[ordered_cols]
            rename_map = {c[0]: c[1] for c in loan_column_config}
            df_loans.rename(columns=rename_map, inplace=True)
            df_loans.to_excel(writer, index=False, sheet_name='Loans')
            ws = writer.sheets['Loans']
            for col_idx, col in enumerate(df_loans.columns, 1):
                max_len = max(len(str(col)), df_loans[col].astype(str).str.len().max() if not df_loans.empty else 0)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
        
        # Users sheet with proper columns
        if users:
            df_users = pd.DataFrame(users)
            existing_cols = [c[0] for c in user_column_config if c[0] in df_users.columns]
            df_users = df_users[existing_cols]
            rename_map = {c[0]: c[1] for c in user_column_config if c[0] in df_users.columns}
            df_users.rename(columns=rename_map, inplace=True)
            df_users.to_excel(writer, index=False, sheet_name='Users')
            ws = writer.sheets['Users']
            for col_idx, col in enumerate(df_users.columns, 1):
                max_len = max(len(str(col)), df_users[col].astype(str).str.len().max() if not df_users.empty else 0)
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
        
        # Field Configs sheet
        if field_configs:
            df_fields = pd.DataFrame(field_configs)
            df_fields.to_excel(writer, index=False, sheet_name='Field_Configurations')
    
    output.seek(0)
    
    filename = f"mhp_fintech_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.post("/import/loans-excel")
async def import_loans_from_excel(file: UploadFile = File(...), current_user: User = Depends(get_current_user)):
    """Import loans from Excel file"""
    check_admin(current_user)
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
    
    try:
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(BytesIO(contents))
        
        # Column mapping (case-insensitive, maps readable headers and common variations)
        column_mapping = {
            'name': 'name',
            'date': 'month',
            'entry date': 'month',
            'created date': 'month',
            'month': 'month',
            'customer name': 'customer_name',
            'customername': 'customer_name',
            'customer': 'customer_name',
            'company name': 'company_name',
            'companyname': 'company_name',
            'company': 'company_name',
            'contact no': 'contact_no',
            'contact': 'contact_no',
            'contact number': 'contact_no',
            'mobile': 'contact_no',
            'phone': 'contact_no',
            'status': 'status',
            'bank': 'bank',
            'sanction amount': 'sanction',
            'sanction': 'sanction',
            'bank sanctioned': 'sanction',
            'disbursed amount': 'disbursed',
            'disbursed': 'disbursed',
            'remark': 'remark',
            'remarks': 'remark',
            'decline reason': 'decline_reason',
            'decline': 'decline_reason',
            'scheme': 'scheme',
            'case from': 'case_from',
            'location': 'location',
            'city': 'location',
            'branch': 'branch',
            'executive name': 'executive_name',
            'executive': 'executive_name',
            'agent name': 'agent_name',
            'agent': 'agent_name',
            'team manager': 'team_manager',
            'manager': 'team_manager',
            'code': 'code',
            'rate': 'rate',
            'roi': 'rate',
            'interest rate': 'rate',
            'pf': 'pf',
            'insurance': 'insurance',
            'tenure (months)': 'tenure',
            'tenure': 'tenure',
            'subvention': 'subvention',
            'brokerage': 'brokerage_subvention',
            'brokerage subvention': 'brokerage_subvention',
            'technical value': 'technical_value',
            'technicalvalue': 'technical_value',
            'technical': 'technical_value',
            'legal status': 'legal_status',
            'legalstatus': 'legal_status',
            'legal': 'legal_status',
            'login date': 'login_date',
            'logindate': 'login_date',
            'amount': 'amount',
            'loan amount': 'amount',
            'disbursed date': 'disbursed_date',
            'disburseddate': 'disbursed_date',
            'disbursal date': 'disbursed_date',
        }
        
        # Normalize column names
        df.columns = df.columns.str.strip().str.lower()
        
        # Rename columns
        for old_col, new_col in column_mapping.items():
            if old_col in df.columns:
                df.rename(columns={old_col: new_col}, inplace=True)
        
        # Required fields (removed month as it can be auto-generated)
        required_fields = ['customer_name', 'status']
        missing_fields = [field for field in required_fields if field not in df.columns]
        
        if missing_fields:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing_fields)}"
            )
        
        # Auto-generate month if not present (use current month)
        current_month = datetime.now().strftime("%b'%y")  # e.g., "Jan'25"
        
        # Import loans with duplicate detection
        imported_count = 0
        skipped_count = 0
        duplicate_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get('customer_name')) or not str(row.get('customer_name')).strip():
                    skipped_count += 1
                    continue
                
                customer_name = str(row.get('customer_name', '')).strip()
                contact_no = str(row.get('contact_no', '')).strip() if pd.notna(row.get('contact_no')) else ''
                bank = str(row.get('bank', '')).strip() if pd.notna(row.get('bank')) else ''
                
                # Duplicate detection: check if same customer_name + contact_no + bank exists
                dup_query = {"customer_name": customer_name}
                if contact_no:
                    dup_query["contact_no"] = contact_no
                if bank:
                    dup_query["bank"] = bank
                
                existing = await db.loan_applications.find_one(dup_query, {"_id": 0, "id": 1})
                if existing:
                    duplicate_count += 1
                    continue
                
                # Prepare loan data
                loan_data = {
                    "id": str(uuid.uuid4()),
                    "customer_name": str(row.get('customer_name', '')).strip(),
                    "company_name": str(row.get('company_name', '')).strip() if pd.notna(row.get('company_name')) else '',
                    "contact_no": str(row.get('contact_no', '')).strip() if pd.notna(row.get('contact_no')) else '',
                    "status": str(row.get('status', 'Pending')).strip(),
                    "bank": str(row.get('bank', '')).strip() if pd.notna(row.get('bank')) else '',
                    "sanction": str(row.get('sanction', '')).strip() if pd.notna(row.get('sanction')) else '',
                    "disbursed": str(row.get('disbursed', '')).strip() if pd.notna(row.get('disbursed')) else '',
                    "remark": str(row.get('remark', '')).strip() if pd.notna(row.get('remark')) else '',
                    "decline_reason": str(row.get('decline_reason', '')).strip() if pd.notna(row.get('decline_reason')) else '',
                    "scheme": str(row.get('scheme', '')).strip() if pd.notna(row.get('scheme')) else '',
                    "case_from": str(row.get('case_from', '')).strip() if pd.notna(row.get('case_from')) else '',
                    "location": str(row.get('location', '')).strip() if pd.notna(row.get('location')) else '',
                    "branch": str(row.get('branch', '')).strip() if pd.notna(row.get('branch')) else '',
                    "executive_name": str(row.get('executive_name', '')).strip() if pd.notna(row.get('executive_name')) else '',
                    "agent_name": str(row.get('agent_name', current_user.name)).strip() if pd.notna(row.get('agent_name')) else current_user.name,
                    "team_manager": str(row.get('team_manager', '')).strip() if pd.notna(row.get('team_manager')) else '',
                    "code": str(row.get('code', '')).strip() if pd.notna(row.get('code')) else '',
                    "rate": str(row.get('rate', '')).strip() if pd.notna(row.get('rate')) else '',
                    "pf": str(row.get('pf', '')).strip() if pd.notna(row.get('pf')) else '',
                    "insurance": str(row.get('insurance', '')).strip() if pd.notna(row.get('insurance')) else '',
                    "tenure": str(row.get('tenure', '')).strip() if pd.notna(row.get('tenure')) else '',
                    "subvention": str(row.get('subvention', '')).strip() if pd.notna(row.get('subvention')) else '',
                    "brokerage_subvention": str(row.get('brokerage_subvention', '')).strip() if pd.notna(row.get('brokerage_subvention')) else '',
                    "month": str(row.get('month', current_month)).strip() if pd.notna(row.get('month')) else current_month,
                    "created_by": current_user.id,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                
                # Insert into database
                await db.loan_applications.insert_one(loan_data)
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                skipped_count += 1
        
        return {
            "message": "Import completed",
            "imported": imported_count,
            "skipped": skipped_count,
            "duplicates": duplicate_count,
            "total_rows": len(df),
            "errors": errors[:10] if errors else []
        }
        
    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

# AI-powered features
from emergentintegrations.llm.chat import LlmChat, UserMessage
import json

class AISearchRequest(BaseModel):
    query: str

class AIAnalysisRequest(BaseModel):
    month: Optional[str] = None
    question: str

@api_router.post("/ai/search")
async def ai_natural_language_search(request: AISearchRequest, current_user: User = Depends(get_current_user)):
    """Convert natural language query to filters"""
    try:
        llm_key = os.environ.get('EMERGENT_LLM_KEY')
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"ai-search-{current_user.id}",
            system_message="""You are an AI assistant that converts natural language queries into structured filters for a loan MIS system.

Available filter fields:
- status: string (e.g., "Pending", "Approved", "Disbursed", "Declined", "Hold")
- bank: string (e.g., "HDFC", "ICICI", "SBI", "Axis")
- month: string (e.g., "Jan'25", "Feb'25")
- location: string (city or state)
- agent_name: string (executive name)
- customer_name: string
- product_type: string (e.g., "Personal Loan", "Business Loan")

Return ONLY a valid JSON object with the filters. Example:
{"status": "Disbursed", "bank": "HDFC", "month": "Jan'25"}

If no specific filters can be extracted, return an empty object: {}"""
        ).with_model("openai", "gpt-4o-mini")
        
        user_message = UserMessage(text=f"Query: {request.query}")
        response = await chat.send_message(user_message)
        
        # Parse JSON response
        try:
            filters = json.loads(response)
        except:
            filters = {}
        
        return {"filters": filters, "query": request.query}
    except Exception as e:
        logger.error(f"AI search error: {str(e)}")
        raise HTTPException(status_code=500, detail="AI search failed")

@api_router.post("/ai/suggestions")
async def ai_smart_suggestions(field: str, partial_value: str, current_user: User = Depends(get_current_user)):
    """Get AI-powered suggestions for field values"""
    try:
        # Get existing values from database
        loans = await db.loan_applications.find({}, {"_id": 0, field: 1}).to_list(100000)
        existing_values = list(set([loan.get(field, "") for loan in loans if loan.get(field)]))
        
        # Filter matching values
        suggestions = [val for val in existing_values if partial_value.lower() in val.lower()][:5]
        
        return {"suggestions": suggestions}
    except Exception as e:
        logger.error(f"Suggestions error: {str(e)}")
        return {"suggestions": []}

@api_router.post("/ai/analyze")
async def ai_data_analysis(request: AIAnalysisRequest, current_user: User = Depends(get_current_user)):
    """AI-powered data analysis and insights"""
    try:
        # Fetch data
        query = {}
        accessible_ids = await get_accessible_user_ids(current_user)
        if accessible_ids is not None:
            query["created_by"] = {"$in": accessible_ids}
        
        if request.month:
            query["month"] = request.month
        
        loans = await db.loan_applications.find(query, {"_id": 0}).to_list(10000)
        
        # Prepare data summary for AI
        total_loans = len(loans)
        status_counts = {}
        bank_counts = {}
        total_sanction = 0
        total_disbursed = 0
        
        for loan in loans:
            status = loan.get('status', 'Unknown')
            bank = loan.get('bank', 'Unknown')
            status_counts[status] = status_counts.get(status, 0) + 1
            bank_counts[bank] = bank_counts.get(bank, 0) + 1
            
            try:
                if loan.get('sanction'):
                    total_sanction += float(str(loan.get('sanction', 0)).replace(',', ''))
            except:
                pass
            try:
                if loan.get('disbursed'):
                    total_disbursed += float(str(loan.get('disbursed', 0)).replace(',', ''))
            except:
                pass
        
        data_summary = {
            "total_loans": total_loans,
            "status_breakdown": status_counts,
            "bank_breakdown": bank_counts,
            "total_sanction_amount": total_sanction,
            "total_disbursed_amount": total_disbursed,
            "month": request.month or "All months"
        }
        
        # Ask AI to analyze
        llm_key = os.environ.get('EMERGENT_LLM_KEY')
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"ai-analysis-{current_user.id}",
            system_message="""You are a financial data analyst AI. Analyze loan data and provide clear, actionable insights.
Be concise and highlight key trends, patterns, and recommendations."""
        ).with_model("openai", "gpt-4o-mini")
        
        prompt = f"""Analyze this loan data and answer: {request.question}

Data Summary:
{json.dumps(data_summary, indent=2)}

Provide a clear, concise analysis with specific numbers and actionable insights."""
        
        user_message = UserMessage(text=prompt)
        response = await chat.send_message(user_message)
        
        return {
            "analysis": response,
            "data_summary": data_summary
        }
    except Exception as e:
        logger.error(f"AI analysis error: {str(e)}")
        raise HTTPException(status_code=500, detail="AI analysis failed")

# ==================== Master File Endpoints ====================

@api_router.get("/master/banks")
async def get_master_banks(current_user: User = Depends(get_current_user)):
    banks = await db.master_banks.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return banks

@api_router.post("/master/banks")
async def add_master_bank(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bank name is required")
    existing = await db.master_banks.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Bank name already exists")
    bank = {
        "id": str(uuid.uuid4()),
        "name": name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    await db.master_banks.insert_one(bank)
    del bank["_id"]
    return bank

@api_router.put("/master/banks/{bank_id}")
async def update_master_bank(bank_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Bank name is required")
    existing = await db.master_banks.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": bank_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Bank name already exists")
    result = await db.master_banks.update_one({"id": bank_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Bank not found")
    return {"id": bank_id, "name": name}

@api_router.delete("/master/banks/{bank_id}")
async def delete_master_bank(bank_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_banks.delete_one({"id": bank_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Bank not found")
    return {"message": "Bank deleted"}

@api_router.get("/master/agents")
async def get_master_agents(current_user: User = Depends(get_current_user)):
    agents = await db.master_agents.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return agents

@api_router.post("/master/agents")
async def add_master_agent(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Agent name is required")
    agent = {
        "id": str(uuid.uuid4()),
        "name": name,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.id
    }
    await db.master_agents.insert_one(agent)
    del agent["_id"]
    return agent

@api_router.put("/master/agents/{agent_id}")
async def update_master_agent(agent_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Agent name is required")
    result = await db.master_agents.update_one({"id": agent_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Agent not found")
    return {"id": agent_id, "name": name}

@api_router.delete("/master/agents/{agent_id}")
async def delete_master_agent(agent_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_agents.delete_one({"id": agent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agent not found")
    return {"message": "Agent deleted"}

# --- Master Companies ---
@api_router.get("/master/companies")
async def get_master_companies(current_user: User = Depends(get_current_user)):
    return await db.master_companies.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api_router.post("/master/companies")
async def add_master_company(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Company name is required")
    existing = await db.master_companies.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Company name already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.id}
    await db.master_companies.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.put("/master/companies/{item_id}")
async def update_master_company(item_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Company name is required")
    existing = await db.master_companies.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": item_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Company name already exists")
    result = await db.master_companies.update_one({"id": item_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"id": item_id, "name": name}

@api_router.delete("/master/companies/{item_id}")
async def delete_master_company(item_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_companies.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Company not found")
    return {"message": "Company deleted"}

# --- Master Branches ---
@api_router.get("/master/branches")
async def get_master_branches(current_user: User = Depends(get_current_user)):
    return await db.master_branches.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api_router.post("/master/branches")
async def add_master_branch(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Branch name is required")
    existing = await db.master_branches.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Branch name already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.id}
    await db.master_branches.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.put("/master/branches/{item_id}")
async def update_master_branch(item_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Branch name is required")
    existing = await db.master_branches.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": item_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Branch name already exists")
    result = await db.master_branches.update_one({"id": item_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"id": item_id, "name": name}

@api_router.delete("/master/branches/{item_id}")
async def delete_master_branch(item_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_branches.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Branch not found")
    return {"message": "Branch deleted"}

# --- Master Locations ---
@api_router.get("/master/locations")
async def get_master_locations(current_user: User = Depends(get_current_user)):
    return await db.master_locations.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api_router.post("/master/locations")
async def add_master_location(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Location name is required")
    existing = await db.master_locations.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Location already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.id}
    await db.master_locations.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.put("/master/locations/{item_id}")
async def update_master_location(item_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Location name is required")
    existing = await db.master_locations.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": item_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Location already exists")
    result = await db.master_locations.update_one({"id": item_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"id": item_id, "name": name}

@api_router.delete("/master/locations/{item_id}")
async def delete_master_location(item_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_locations.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Location not found")
    return {"message": "Location deleted"}

# --- Master Categories ---
@api_router.get("/master/categories")
async def get_master_categories(current_user: User = Depends(get_current_user)):
    return await db.master_categories.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api_router.post("/master/categories")
async def add_master_category(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")
    existing = await db.master_categories.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.id}
    await db.master_categories.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.put("/master/categories/{item_id}")
async def update_master_category(item_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Category name is required")
    existing = await db.master_categories.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": item_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Category already exists")
    result = await db.master_categories.update_one({"id": item_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"id": item_id, "name": name}

@api_router.delete("/master/categories/{item_id}")
async def delete_master_category(item_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_categories.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted"}

# --- Master Products ---
@api_router.get("/master/products")
async def get_master_products(current_user: User = Depends(get_current_user)):
    return await db.master_products.find({}, {"_id": 0}).sort("name", 1).to_list(1000)

@api_router.post("/master/products")
async def add_master_product(request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Product name is required")
    existing = await db.master_products.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=400, detail="Product already exists")
    doc = {"id": str(uuid.uuid4()), "name": name, "created_at": datetime.now(timezone.utc).isoformat(), "created_by": current_user.id}
    await db.master_products.insert_one(doc)
    del doc["_id"]
    return doc

@api_router.put("/master/products/{item_id}")
async def update_master_product(item_id: str, request: Request, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    data = await request.json()
    name = data.get("name", "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Product name is required")
    existing = await db.master_products.find_one({"name": {"$regex": f"^{name}$", "$options": "i"}, "id": {"$ne": item_id}})
    if existing:
        raise HTTPException(status_code=400, detail="Product already exists")
    result = await db.master_products.update_one({"id": item_id}, {"$set": {"name": name, "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"id": item_id, "name": name}

@api_router.delete("/master/products/{item_id}")
async def delete_master_product(item_id: str, current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    result = await db.master_products.delete_one({"id": item_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Product not found")
    return {"message": "Product deleted"}

# ==================== DB Backup Endpoints ====================

@api_router.get("/backup/download")
async def download_backup(current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    import json as json_lib
    backup = {}
    collections = ["users", "loan_applications", "schemes", "statuses", "master_banks", "master_agents", "master_companies", "master_branches", "master_locations", "master_categories", "master_products", "agent_targets"]
    for col_name in collections:
        docs = await db[col_name].find({}, {"_id": 0}).to_list(100000)
        backup[col_name] = docs
    backup["metadata"] = {
        "backup_date": datetime.now(timezone.utc).isoformat(),
        "backed_up_by": current_user.email,
        "collections": {col: len(backup[col]) for col in collections}
    }
    content = json_lib.dumps(backup, indent=2, default=str)
    filename = f"mhp_backup_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
    return Response(
        content=content,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/backup/stats")
async def get_backup_stats(current_user: User = Depends(get_current_user)):
    check_admin(current_user)
    collections = ["users", "loan_applications", "schemes", "statuses", "master_banks", "master_agents", "master_companies", "master_branches", "master_locations", "master_categories", "master_products", "agent_targets"]
    stats = {}
    total = 0
    for col_name in collections:
        count = await db[col_name].count_documents({})
        stats[col_name] = count
        total += count
    stats["total_records"] = total
    return stats

@api_router.post("/backup/import")
async def import_backup(file: UploadFile = File(...), mode: str = Form("merge"), current_user: User = Depends(get_current_user)):
    """Import a JSON backup file. mode='merge' adds new records, mode='replace' wipes and replaces."""
    check_admin(current_user)
    import json as json_lib
    try:
        content = await file.read()
        data = json_lib.loads(content)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON file")

    collections = ["users", "loan_applications", "schemes", "statuses", "master_banks", "master_agents", "master_companies", "master_branches", "master_locations", "master_categories", "master_products", "agent_targets"]
    results = {}
    total_imported = 0

    for col_name in collections:
        if col_name not in data:
            continue
        records = data[col_name]
        if not isinstance(records, list) or len(records) == 0:
            continue

        if mode == "replace":
            await db[col_name].delete_many({})
            if records:
                for rec in records:
                    rec.pop("_id", None)
                await db[col_name].insert_many(records)
            results[col_name] = {"action": "replaced", "count": len(records)}
            total_imported += len(records)
        else:
            inserted = 0
            skipped = 0
            for rec in records:
                rec.pop("_id", None)
                rec_id = rec.get("id")
                rec_name = rec.get("name")
                if rec_id:
                    exists = await db[col_name].find_one({"id": rec_id})
                    if exists:
                        skipped += 1
                        continue
                elif rec_name and col_name.startswith("master_"):
                    exists = await db[col_name].find_one({"name": rec_name})
                    if exists:
                        skipped += 1
                        continue
                await db[col_name].insert_one(rec)
                inserted += 1
            results[col_name] = {"action": "merged", "inserted": inserted, "skipped": skipped}
            total_imported += inserted

    return {"message": f"Import complete. {total_imported} records imported.", "details": results, "total_imported": total_imported}



# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def create_default_admin():
    """Create default admin user and schemes if they don't exist"""
    try:
        admin_email = "admin@mhpfintech.com"
        admin_password = "Admin@123"
        
        # Check if admin exists
        existing_admin = await db.users.find_one({"email": admin_email})
        
        if not existing_admin:
            # Create admin user
            admin_user = {
                "id": str(uuid.uuid4()),
                "email": admin_email,
                "name": "Admin User",
                "role": "admin",
                "team_code": None,
                "manager_id": None,
                "password": hash_password(admin_password),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.users.insert_one(admin_user)
            logger.info(f"✅ Created default admin user: {admin_email}")
        else:
            # Ensure existing admin has admin role
            if existing_admin.get('role') != 'admin':
                await db.users.update_one(
                    {"email": admin_email},
                    {"$set": {"role": "admin"}}
                )
                logger.info(f"✅ Updated role to 'admin' for {admin_email}")
            else:
                logger.info(f"✅ Admin user already exists: {admin_email}")
        
        # Create default manager user
        manager_email = "manager@mhpfintech.com"
        manager_password = "Admin@123"
        existing_manager = await db.users.find_one({"email": manager_email})
        
        if not existing_manager:
            manager_user = {
                "id": str(uuid.uuid4()),
                "email": manager_email,
                "name": "Manager User",
                "role": "manager",
                "team_code": "TEAM-A",
                "manager_id": None,
                "active": True,
                "password": hash_password(manager_password),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(manager_user)
            logger.info(f"✅ Created default manager user: {manager_email}")
            existing_manager = manager_user
        else:
            # Ensure existing manager has correct role
            if existing_manager.get('role') != 'manager':
                await db.users.update_one({"email": manager_email}, {"$set": {"role": "manager", "team_code": "TEAM-A"}})
            logger.info(f"✅ Manager user already exists: {manager_email}")
        
        # Create default agent user and assign to manager
        agent_email = "agent@mhpfintech.com"
        agent_password = "Admin@123"
        manager_id = existing_manager.get('id')
        existing_agent = await db.users.find_one({"email": agent_email})
        
        if not existing_agent:
            agent_user = {
                "id": str(uuid.uuid4()),
                "email": agent_email,
                "name": "Agent User",
                "role": "agent",
                "team_code": "TEAM-A",
                "manager_id": manager_id,
                "active": True,
                "password": hash_password(agent_password),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(agent_user)
            logger.info(f"✅ Created default agent user: {agent_email} (assigned to manager)")
        else:
            # Ensure agent is assigned to manager and has valid id
            updates = {}
            if not existing_agent.get('id'):
                updates["id"] = str(uuid.uuid4())
            if existing_agent.get('manager_id') != manager_id:
                updates["manager_id"] = manager_id
            if existing_agent.get('team_code') != "TEAM-A":
                updates["team_code"] = "TEAM-A"
            if existing_agent.get('role') != 'agent':
                updates["role"] = "agent"
            if updates:
                await db.users.update_one({"email": agent_email}, {"$set": updates})
                logger.info(f"✅ Updated agent {agent_email}: {list(updates.keys())}")
            else:
                logger.info(f"✅ Agent user already exists and assigned: {agent_email}")
        
        # Initialize default schemes ONLY if collection is completely empty (first run)
        scheme_count = await db.schemes.count_documents({})
        if scheme_count == 0:
            default_schemes = [
                {"name": "GST", "description": "GST Loan Scheme"},
                {"name": "Degree Surrogate", "description": "Degree Surrogate Loan"},
                {"name": "Prime Banking", "description": "Prime Banking Loan"},
                {"name": "LTBL", "description": "LTBL Loan Scheme"},
                {"name": "SEPL Income Programme", "description": "SEPL Income Programme"},
                {"name": "Car loan", "description": "Car Loan Scheme"},
                {"name": "Education Loan", "description": "Education Loan Scheme"}
            ]
            
            for scheme_data in default_schemes:
                scheme = {
                    "id": str(uuid.uuid4()),
                    "name": scheme_data["name"],
                    "description": scheme_data["description"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                }
                await db.schemes.insert_one(scheme)
                logger.info(f"Created default scheme: {scheme_data['name']}")
        
        scheme_count = await db.schemes.count_documents({})
        logger.info(f"Total schemes in database: {scheme_count}")
        
        # Initialize default statuses ONLY if collection is completely empty (first run)
        status_count = await db.statuses.count_documents({})
        if status_count == 0:
            default_statuses = [
                {"name": "Pending", "description": "Application pending", "color": "#FFA500", "order": 1},
                {"name": "Login", "description": "Login stage", "color": "#00BFFF", "order": 2},
                {"name": "Query", "description": "Query raised", "color": "#FFD700", "order": 3},
                {"name": "Approved", "description": "Application approved", "color": "#32CD32", "order": 4},
                {"name": "Post PD Docs", "description": "Post PD documents stage", "color": "#9370DB", "order": 5},
                {"name": "Sanctioned", "description": "Loan sanctioned", "color": "#00CED1", "order": 6},
                {"name": "Disbursed", "description": "Loan disbursed", "color": "#228B22", "order": 7},
                {"name": "Decline", "description": "Application declined", "color": "#DC143C", "order": 8},
                {"name": "Hold", "description": "Application on hold", "color": "#FF8C00", "order": 9},
                {"name": "Rejected", "description": "Application rejected", "color": "#B22222", "order": 10}
            ]
            
            for status_data in default_statuses:
                status = {
                    "id": str(uuid.uuid4()),
                    "name": status_data["name"],
                    "description": status_data["description"],
                    "color": status_data["color"],
                    "order": status_data["order"],
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "created_by": admin_id
                }
                await db.statuses.insert_one(status)
                logger.info(f"Created default status: {status_data['name']}")
        
        status_count = await db.statuses.count_documents({})
        logger.info(f"Total statuses in database: {status_count}")
        
        # Create database indexes for performance
        await db.loan_applications.create_index("created_by")
        await db.loan_applications.create_index("month")
        await db.loan_applications.create_index("status")
        await db.loan_applications.create_index("bank")
        await db.loan_applications.create_index("created_at")
        await db.loan_applications.create_index([("customer_name", 1), ("contact_no", 1), ("bank", 1)])
        await db.users.create_index("email", unique=True)
        await db.master_banks.create_index("name", unique=True)
        await db.master_agents.create_index("name")
        await db.master_companies.create_index("name", unique=True)
        await db.master_branches.create_index("name", unique=True)
        await db.master_locations.create_index("name", unique=True)
        await db.master_categories.create_index("name", unique=True)
        await db.master_products.create_index("name", unique=True)
        logger.info("Database indexes created")
        
    except Exception as e:
        logger.error(f"❌ Error in startup: {str(e)}")

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()